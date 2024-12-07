import streamlit as st
import openpyxl
from io import BytesIO
import uuid

# Função que salva os dados no Excel
from openpyxl import load_workbook, Workbook

def salvar_dados_excel(segmento, funcionarios, anos_operando):
    try:
        wb = load_workbook('dados_coletados.xlsx')  # Tenta abrir o arquivo existente
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()  # Se o arquivo não existe, cria um novo
        ws = wb.active
        ws.append(["Segmento de Mercado", "Número de Funcionários", "Anos Operando"])  # Cabeçalhos

    ws.append([segmento, funcionarios, anos_operando])  # Adiciona os dados
    wb.save('dados_coletados.xlsx')  # Salva o arquivo


def atualizar_fluxo_caixa(segmento, funcionarios, anos_operando, codigo_unico):
    # Carregar o modelo de fluxo de caixa
    fluxo_caixa = openpyxl.load_workbook("Fluxo_de_Caixa.xlsx")
    ws = fluxo_caixa.active

    # Atualizar as células com os dados do chatbot
    ws['A1'] = segmento
    ws['A2'] = funcionarios
    ws['A3'] = anos_operando

    # Salvar a planilha em memória (não no disco)
    output = BytesIO()
    fluxo_caixa.save(output)
    output.seek(0)  # Garantir que o ponteiro esteja no início para leitura

    return output

def chatbot():
    st.title("Coleta de Dados - SaaS")
    st.write("Por favor, responda às perguntas abaixo:")

    # Perguntas para o usuário
    segmento = st.text_input("Qual é o seu segmento de mercado?")
    funcionarios = st.number_input("Quantos funcionários você tem?", step=1, min_value=0)
    anos_operando = st.number_input("Há quantos anos sua empresa está operando?", step=1, min_value=0)

    # Botão para enviar os dados
    if st.button("Enviar"):
        if segmento and funcionarios and anos_operando:
            # Gerar o código único
            codigo_unico = str(uuid.uuid4())

            # Atualiza a planilha com os dados do chatbot
            fluxo_caixa_em_memoria = atualizar_fluxo_caixa(segmento, funcionarios, anos_operando, codigo_unico)

            # Exibir um resumo das informações
            st.success("Dados enviados com sucesso!")
            st.write(f"- Segmento de Mercado: {segmento}")
            st.write(f"- Número de Funcionários: {funcionarios}")
            st.write(f"- Anos Operando: {anos_operando}")
            
            # Criar o botão de download para o arquivo gerado em memória
            st.download_button(
                label="Baixar Fluxo de Caixa Atualizado",
                data=fluxo_caixa_em_memoria,
                file_name=f"Fluxo_de_Caixa_{codigo_unico}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
	# Salva os dados no Excel
            salvar_dados_excel(segmento, funcionarios, anos_operando)

        else:
            st.error("Por favor, preencha todas as informações!")

# Executa o chatbot
if __name__ == "__main__":
    chatbot()