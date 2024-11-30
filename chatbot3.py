import streamlit as st
from openpyxl import Workbook, load_workbook
import os
import requests

# Função para criar uma pasta temporária se não existir
def criar_pasta_temp():
    os.makedirs("temp", exist_ok=True)

# Função para fazer o download do arquivo Fluxo_de_Caixa.xlsx do GitHub
def download_arquivo_fluxo():
    # URL do arquivo no GitHub (ajuste se necessário)
    url = "https://github.com/edugr844/obeze/raw/main/Fluxo_de_Caixa.xlsx"
    caminho_temp = os.path.join("temp", "Fluxo_de_Caixa.xlsx")
    
    # Fazendo o download do arquivo
    response = requests.get(url)
    if response.status_code != 200:
        raise ValueError("Erro ao baixar o arquivo do GitHub. Verifique a URL.")
    
    # Salvando o arquivo localmente
    with open(caminho_temp, "wb") as f:
        f.write(response.content)
    
    return caminho_temp

# Função para salvar os dados no arquivo Fluxo de Caixa
def salvar_fluxo_de_caixa(segmento, funcionarios, anos_operando):
    criar_pasta_temp()  # Garantir que a pasta existe
    
    # Baixa o arquivo original do GitHub
    caminho_fluxo_original = download_arquivo_fluxo()
    wb = load_workbook(caminho_fluxo_original)
    ws = wb.active

    # Atualiza as células com os dados fornecidos
    ws['A1'] = f"Segmento: {segmento}"
    ws['A2'] = f"Funcionários: {funcionarios}"
    ws['A3'] = f"Anos Operando: {anos_operando}"

    # Nome do arquivo final
    file_name = f"fluxo_de_caixa_{segmento.replace(' ', '_')}.xlsx"
    temp_file_path = os.path.join("temp", file_name)
    wb.save(temp_file_path)

    return temp_file_path

# Função para salvar os dados coletados em um novo arquivo
def salvar_dados_coletados(segmento, funcionarios, anos_operando):
    criar_pasta_temp()  # Garantir que a pasta existe

    # Cria uma nova planilha e adiciona os dados
    wb = Workbook()
    ws = wb.active
    ws.append(["Segmento de Mercado", "Número de Funcionários", "Anos Operando"])
    ws.append([segmento, funcionarios, anos_operando])

    # Nome do arquivo final
    file_name = f"dados_coletados_{segmento.replace(' ', '_')}.xlsx"
    temp_file_path = os.path.join("temp", file_name)
    wb.save(temp_file_path)

    return temp_file_path

# Função principal do chatbot
def chatbot():
    st.title("Obezê")
    st.write("Por favor, responda às perguntas abaixo:")

    # Perguntas para o usuário
    segmento = st.text_input("Qual é o seu segmento de mercado?")
    funcionarios = st.number_input("Quantos funcionários você tem?", step=1, min_value=0)
    anos_operando = st.number_input("Há quantos anos sua empresa está operando?", step=1, min_value=0)

    # Botão para enviar os dados
    if st.button("Enviar"):
        if segmento and funcionarios and anos_operando:
            st.success("Dados enviados com sucesso!")
            st.write("Resumo das informações coletadas:")
            st.write(f"- Segmento de Mercado: {segmento}")
            st.write(f"- Número de Funcionários: {funcionarios}")
            st.write(f"- Anos Operando: {anos_operando}")
            
            try:
                # Salva o Fluxo de Caixa atualizado
                fluxo_file = salvar_fluxo_de_caixa(segmento, funcionarios, anos_operando)

                # Salva os dados coletados em um novo arquivo
                dados_file = salvar_dados_coletados(segmento, funcionarios, anos_operando)

                # Botões de download
                with open(fluxo_file, "rb") as f_fluxo:
                    st.download_button(
                        label="Baixar Fluxo de Caixa Atualizado",
                        data=f_fluxo,
                        file_name=os.path.basename(fluxo_file),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                with open(dados_file, "rb") as f_dados:
                    st.download_button(
                        label="Baixar Dados Coletados",
                        data=f_dados,
                        file_name=os.path.basename(dados_file),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Ocorreu um erro: {e}")
        else:
            st.error("Por favor, preencha todas as informações!")

# Executa o chatbot
if __name__ == "__main__":
    chatbot()
